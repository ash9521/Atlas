from __future__ import annotations

from openpyxl import load_workbook

from atlas.discovery.connectors import BaseConnector
from atlas.discovery.models import InputSource, Observation


class ExcelConnector(BaseConnector):
    @property
    def name(self) -> str:
        return "excel"

    def acquire(
        self,
        source: InputSource,
    ) -> list[Observation]:
        workbook = load_workbook(
            filename=source.location,
            read_only=True,
            data_only=True,
        )

        try:
            worksheet = workbook.active

            if worksheet is None:
                return []

            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                return []

            headers = [
                str(value).strip() if value is not None else ""
                for value in rows[0]
            ]

            observations: list[Observation] = []

            for row in rows[1:]:
                payload = {
                    headers[index]: value
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }

                observations.append(
                    Observation.create(
                        connector=self.name,
                        payload=payload,
                        metadata={
                            "worksheet": worksheet.title,
                        },
                    ),
                )

            return observations

        finally:
            workbook.close()
